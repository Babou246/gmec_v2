import random
import string
from datetime import datetime, date
# import date
import os
# from flask_sqlalchemy import SQLAlchemy
from models import User,app,TypeDefaut,db,Role
from flask import Flask, render_template, request, redirect, url_for, flash,session
# from flask_login import login_user
from flask_login import LoginManager,UserMixin,login_user,login_required,logout_user,current_user
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import bcrypt
import glob
from models import User,db,app,TypeDefaut,Role,Service,Fichier,Fichier_charger,Ticket,Corbeille
from flask_migrate import Migrate
from decimal import Decimal
import openpyxl
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
import hashlib
# import mail
from os.path import join, dirname, realpath
from flask_bcrypt import check_password_hash, generate_password_hash,Bcrypt
from openpyxl import load_workbook
import pandas as pd
from sqlalchemy import create_engine
import csv
import numpy as np
# from utils import *
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired
from mail_rappel import envoyer_rappel,envoyer_mail_rappel
from flask_paginate import Pagination, get_page_parameter

# Configuration 

login_manager=LoginManager()
login_manager.init_app(app)
login_manager.login_view='login'
migrate = Migrate(app, db)
mail = Mail(app)
bcrypt = Bcrypt()


app.config['MAIL_SERVER'] = 'smtp.hushmail.com'
app.config['MAIL_USERNAME'] = 'babou@gmail.com'
app.config['MAIL_PASSWORD'] = ''
app.config['MAIL_DEBUG'] = True
app.config['MAIL_DEFAULT_SENDER'] = 'babou@gmail.com'
app.config['MAIL_PORT'] =12458
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USE_TLS'] = False
mail = Mail(app)


with app.app_context(): 
    db.create_all()

                                                    ########################################################
                                                    #                     Utilitaire                       #
                                                    ######################################################## 

def send_notification_emails(agents):
    for agent in agents:
        recipient = agent['email']
        supervisor = agent['supervisor']
        subject = 'Notification de chargement'
        body = f"Bonjour {agent['prenom']} {agent['nom']}," \
               f"\nLe chargement des échantillons de contrôle des défauts du mois de AAAA/MM est terminé." \
               f"Nous vous invitons à vous connecter à QUALITE pour traiter/commenter les erreurs critiques vous concernant." \
               f"\nCordialement," \
               f"\nL'équipe QUALITE"
        send_email(recipient, supervisor, subject, body)


def send_notification_email(login, password):
    msg = Message('Notification de création d\'utilisateur',
                  sender='babou@sonatel-orange.com',
                  recipients=[login])
    msg.body = f'Vos identifiants de connexion : \nLogin : {login}\nMot de passe : {password}'
    return mail.send(msg)


def generate_code():
    dernier_code = get_last_code_from_database() or 0
    # Générer le nouveau code en incrémentant le dernier code utilisé
    nouveau_code = dernier_code + 1
    return nouveau_code


def enregistrer_defauts(liste_defauts):
    # Convertir la liste des défauts en un format adapté pour l'enregistrement dans la base de données
    defauts_enregistrement = []
    for defaut in liste_defauts:
        defaut_enregistrement = {
            'code': defaut['code'],
            'type': defaut['type'],
            'description': defaut['description'],
            'date_debut': defaut['date_debut'],
            'date_fin': defaut['date_fin']
        }
        defauts_enregistrement.append(defaut_enregistrement)
    
    # Enregistrer les défauts dans la base de données
    save_defauts_to_database(defauts_enregistrement)

def save_defauts_to_database(defauts):
    for defaut in defauts:
        defaut_enregistre = Defaut(
            code=defaut['code'],
            type=defaut['type'],
            description=defaut['description'],
            date_debut=defaut['date_debut'],
            date_fin=defaut['date_fin']
        )
        db.session.add(defaut_enregistre)
    db.session.commit()


ALLOWED_EXTENSIONS = {'csv', 'xlsx'}
app.config['UPLOAD_FOLDER'] = 'uploads'
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, user_id)


@app.route("/logout")
def logout():
    logout_user()
    return redirect('login')


# ########## Tester mon mail
@app.route('/sender')
def sender():
    msg = Message('Bonjour Babou',recipients=['giyel66581@aramask.com'])
    mail.send(msg)

    return 'sent successfully'

                        ########################################################
                        #                     Dashboard                        #
                        ########################################################


@app.route("/sonatel-gmec/menu", methods=('GET', 'POST'))
@login_required
def home():
    print('cureent_user ====>',current_user.service.nom)
    return render_template('pages/menu.html')


                        ########################################################
                        #                     Profil                           #
                        ########################################################


@app.route("/sonatel-gmec/profils", methods=['POST','GET'])
@login_required
def profils():
    dim = date.today() 
    date_saisi= dim.strftime('%d-%m-%Y') 
    role = Role.query.all()
    return render_template('pages/profils.html', role=role) 


@app.route('/sonatel-gmec/profile_modif/<string:id>', methods=['POST'])
@login_required
def profile_modif(id):
    user = User.query.get(id)

    if not user:
        flash('Utilisateur non trouvé.')
        return redirect(url_for('monprofil'))

    if request.method == 'POST':
        user.email = request.form.get('email')
        user.login = request.form.get('login')
        user.password = request.form.get('password')

        db.session.commit()
        flash('Utilisateur modifié avec succès.', 'success')

    return redirect(url_for('monprofil'))


@app.route("/sonatel-gmec/monprofil")
def monprofil():
    print('les sessions id',session)
    user = current_user
    return render_template('pages/monprofil.html',user=user)


                        ########################################################
                        #                     Athentication                    #
                        ########################################################


@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    # print('===========>',current_user,session)
    if current_user.is_authenticated:
        return redirect(url_for('home'))

    if request.method == 'POST':
        login = request.form.get('login')
        password = request.form.get('password')
        # print('===========>',current_user.nom)

        user = User.query.filter_by(login=login).first()
        # Vérifier le mot de passe
        if user.password == password:
            login_user(user)
            return redirect(url_for('home'))
        else:
            flash('Mot de passe incorrect', 'error')
            return redirect(url_for('login'))

    return render_template('pages/login.html')



                            ########################################################
                            #                     Ajouter un User                  #
                            ########################################################


@app.route('/add_user', methods=['GET','POST'])
@login_required
def resolution_utilisateurs():
    if request.method == 'POST':
        prenom = request.form['prenom']
        role = int(request.form['roleid'])
        sigle_service = request.form['sigle_service']
        login = request.form['login']

        nom_abrege = sigle_service + '_' + prenom.replace(' ', '')

        password = "Son@tel2021"
        print('ooooooooo',password)

        # Vérifier si l'utilisateur existe déjà dans la base de données
        existing_user = User.query.filter_by(login=login).first()
        if existing_user:
            flash("L'utilisateur avec le login {} existe déjà.".format(login))

        # Générer le hash du mot de passe
        role = Role.query.get(role)  # Récupérer l'instance de la classe Role avec l'ID de rôle

        user = User(matricule=request.form['matricule'], login=request.form['login'], prenom=request.form['prenom'], nom=request.form['nom'], role=role,
                    sigle_service=request.form['sigle_service'], service_id=int(request.form['service_id']), state=request.form['statut'], email=request.form['email'], nom_abrege=nom_abrege, date_debut=datetime.now(), password="Son@tel2021")

        db.session.add(user)
        db.session.commit()
        # s = URLSafeTimedSerializer('Thisisasecret!')
        # token = s.dumps('becon58494@bodeem.com', salt='email-confirm')
        # msg = Message('Confirm Email', sender='becon58494@bodeem.com', recipients=['becon58494@bodeem.com'])
        # # link = url_for('confirm_email', token=token, _external=True)
        # msg.body = f'Vos identifiants de connexion : \nLogin : {login}\nMot de passe : {password}\n Cliqué sur le lien'
        # mail.send(msg)
        # return redirect(url_for('login'))
    return render_template('pages/menu.html')


@app.route('/confirm_email/<token>')
def confirm_email(token):
    try:
        email = s.loads(token, salt='email-confirm', max_age=3600)
    except SignatureExpired:
        return flash('<h1>The token is expired!</h1>')
    return redirect(url_for('login'))


                            ########################################################
                            #                     Modif PWD                        #
                            ########################################################


@app.route('/changepassword', methods=['GET', 'POST'])
def changepassword():
    if request.method == 'POST':
        login = request.form['login']
        ancien_mot_de_passe = request.form['ancien']
        nouveau_mot_de_passe = request.form['new']
        confirmer_mot_de_passe = request.form['conf']

        # Récupérer l'utilisateur courant
        utilisateur = User.query.filter(User.login == login).first()

        print('=============>',utilisateur)

        # Vérifier si l'ancien mot de passe est correct

        if utilisateur:
            if utilisateur.password != ancien_mot_de_passe:
                flash('Ancien mot de passe incorrect', 'danger')
                return redirect(url_for('changepassword'))

            if len(nouveau_mot_de_passe) > 5:
                # Mettre à jour le mot de passe de l'utilisateur
                utilisateur.password = nouveau_mot_de_passe
                # Vérifier la confirmation du nouveau mot de passe
                if nouveau_mot_de_passe == confirmer_mot_de_passe:
                    db.session.commit()
                    flash('Mot de passe modifié avec succès', 'success')
                    return redirect(url_for('changepassword'))
            else:
                flash('Le mot de passe doit comporter au moins 6 caractères', 'warning')
                return redirect(url_for('changepassword'))

        else:
            flash(f"L'utilisateur {login} n'existe pas dans la base")

    return render_template('pages/login.html')




                                ########################################################
                                #                     Editer PWD                       #
                                ########################################################


@app.route('/modifier_utilisateur/<int:user_id>', methods=['GET', 'POST'])
@login_required
def modifier_utilisateur(user_id):
    user = User.query.get(user_id)
    if not user:
        flash('Utilisateur non trouvé.', 'danger')
        return redirect(url_for('users'))

    if request.method == 'POST':
        user.prenom = request.form.get('prenom')
        user.nom = request.form.get('nom')
        user.role = Role.query.get(int(request.form.get('role')))
        user.service = Service.query.get(int(request.form.get('service')))
        user.sigle_service = request.form.get('sigle_service')
        user.matricule = request.form.get('matricule')
        user.state= request.form.get('statut')
        user.nom_abrege = user.sigle_service+'_'+user.prenom
        if user.state == "Clocturé":
            user.date_fin = datetime.now()

        db.session.commit()
        flash('Utilisateur modifié avec succès.', 'success')
        return redirect(url_for('users'))

    return render_template('pages/utilisateurs.html', user=user)


@app.route('/delete/<string:user_id>', methods=['GET', 'POST'])
@login_required
def delete(user_id):
    user = User.query.get(user_id)
    corebeille = Corbeille.query.get(user_id)

    if user:
        corebeille = Corbeille(
            matricule=user.matricule,
            login=user.login,
            prenom=user.prenom,
            nom=user.nom,
            role = user.role,
            sigle_service=user.sigle_service,
            service_id=user.service.id,
            role_id=user.role.id,
            state="Clocturé",
            email=user.email,
            nom_abrege=user.nom_abrege,
            date_debut=user.date_debut,
            date_fin=datetime.now(),
            password=user.password
        )
        db.session.add(corebeille)
        db.session.delete(user)
        db.session.commit()
        flash('Utilisateur supprimé avec succès.', 'success')
    else:
        flash('Utilisateur introuvable.', 'error')

    return render_template('pages/utilisateurs.html')

@app.route('/delete_dans_corbeille/<int:user_id>')
def delete_dans_cor(user_id):
    corebeille = Corbeille.query.get(user_id)

    if corebeille:
        db.session.delete(corebeille)
        db.session.commit()
    return render_template('corbeille.html')


@app.route('/corbeille')
def get_corbeille():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10  # Nombre de lignes par page

    # Obtenir la liste paginée des utilisateurs
    corbeille_pagination = Corbeille.query.join(Role).join(Service).paginate(page=page, per_page=per_page)
    return render_template('corbeille.html',corbeille_pagination=corbeille_pagination)



@app.route("/sonatel-sovar/guide-utilisateur", methods=['POST','GET'])
def guide():
    return render_template('pages/faq.html') 



                                ########################################################
                                #                     service                          #
                                ########################################################


@app.route("/sonatel-gmec/services", methods=['POST','GET'])
@login_required
def services():
    dim = date.today() 
    date_saisi= dim.strftime('%d-%m-%Y')
    services = Service.query.all()
    user = User.query.all()
    return render_template('pages/services.html',services=services,user=user) 


@app.route("/sonatel-gmec/utilisateurs", methods=['POST','GET'])
@login_required
def users():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10  # Nombre de lignes par page

    # Obtenir la liste paginée des utilisateurs
    users_pagination = User.query.join(Role).join(Service).paginate(page=page, per_page=per_page)
    return render_template('pages/utilisateurs.html', users=users,users_pagination=users_pagination) 


@app.route('/sonatel-gmec/service_users')
@login_required
def service_users():
    users_service = User.query.filter_by(service=current_user.service).all()
    users_services=User.query.all()
    # flash('blabla','success')
    return render_template('service_users.html', users_service=users_service,users_services=users_services)


@app.route('/consulter_services/<int:id>')
@login_required
def consulter_services(id):
    service = Service.query.get(id)
    users_service = User.query.filter(User.service_id == id).all()

    # consulte = Service.query.filter_by(id=service).first()

    return render_template('service_users.html', users_service=users_service)



                                ########################################################
                                #                     Type - Defaut                    #
                                ########################################################


@app.route('/parametrage-defauts', methods=['GET', 'POST'])
@login_required
def parametrage_defauts():
    # print('OOOOOOOOOOOOOOOOOOOO',user_session)
    user = User.query.all()
    if request.method == 'POST':
        if TypeDefaut.query.count() == 0:
            print('OOJHHGGGGGHHHH++++>')
            code = "code_124"
        else:
            print("CCCCCCCCCCCCCCCCCCCCCC")
            code = TypeDefaut.get_next_code()

        user_session = session['_user_id']
        type_defaut = request.form.get('type_defaut')
        description_defaut = request.form.get('description_defaut')
        confirm = request.form.get('oui')
        date_debut = request.form.get('date_debut')
        date_fin = None  # La date de fin est initialisée à None
        commentaires = request.form.get('commentaires')
        validation = request.form.get('validation')
        email_concerne = request.form.get('email')
        evaluer = request.form.get('evaluer')
        n1 = request.form.get('n1')

        print('confirmation >>>>>',confirm)
        listes = []
        for user in user:
            if email_concerne == user.email:
                user_id = user.email
                service= user.service.nom
                # listes.append({'code':code, 'description':description_defaut,'type_default':type_defaut,'commentaires':commentaires,'validation':validation,'service':service})
                # print("PPPPPPPPPPPPPPPPPPPPP==> ",listes)

            else:
                flash(f"Le mail {email_concerne} est insdisponible")

        if type_defaut == "" or description_defaut == "" or date_debut == None:
            flash('Les champs ne doivent pas être vides', 'danger')
            return redirect(url_for('parametrage_defauts'))

        print("Insertion passé avec succés: com et com_n+1",evaluer,n1)

        new_defaut = TypeDefaut(
            code = code,
            type_defaut=type_defaut,
            description_defaut=description_defaut,
            confirm=confirm,
            date_debut=date_debut,
            date_fin=date_fin,
            user_id=user_id,
            commentaires=commentaires,
            validation=validation,
            service=service,
            commentaires_evaluer=evaluer,
            commentaires_n1=n1
        )
        # Ajouter le nouveau défaut à la base de données blabla
        db.session.add(new_defaut)
        db.session.commit()
    defauts = TypeDefaut.query.all()
    elements = TypeDefaut.query.filter(TypeDefaut.service == current_user.service.nom).distinct().order_by(TypeDefaut.date_debut.desc()).all()

    return render_template('parametrage_defauts.html', defauts=defauts,user=user,elements=elements)


@app.route('/modifier-defaut/<int:defaut_id>', methods=['GET', 'POST'])
@login_required
def modifier_defaut(defaut_id):
    
    defaut = TypeDefaut.query.get(defaut_id)

    if request.method == 'POST':
        confirm = request.form.get('OUI')

        print(">>>>>>>>>>", confirm)
        if confirm:
            db.session.commit()
        elif request.form.get('description_defaut') or request.form.get('date_fin') or request.form.get('type_defaut') or request.form.get('commentaires') or request.form.get('validation') or request.form.get('evaluer') or request.form.get('n1'):
            defaut.description_defaut = request.form.get('description_defaut')
            defaut.date_fin = request.form.get('date_fin')
            defaut.type_defaut = request.form.get('type_defaut')
            defaut.commentaires = request.form.get('commentaires')
            defaut.validation = request.form.get('validation')
            evaluer = request.form.get('evaluer')
            n1 = request.form.get('n1')
            # print('===>>>>>',session.get(TypeDefaut, defaut_id),'Object: ',defaut)
            print('===>>>>>',session.get(TypeDefaut, defaut_id))

            defaut.commentaires_evaluer = evaluer
            defaut.commentaires_n1 = n1

            db.session.commit()
        flash('La description du défaut a été modifiée avec succès!', 'success')
        return redirect(url_for('parametrage_defauts'))

    return render_template('modifier_defaut.html', defaut=defaut)


# Interface de chargement des fichiers plats « Défauts »
@app.route('/chargement-defauts', methods=['GET', 'POST'])
@login_required
def chargement_defauts():
    fichiers = Fichier.query.all()
    # print(fichiers)
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Aucun fichier sélectionné.', 'error')
            return redirect(url_for('chargement_defauts'))
        
        file = request.files['file']

        print('============>', file.filename,current_user.id)

        if file.filename == '':
            flash('Aucun fichier sélectionné.', 'error')
            return redirect(url_for('chargement_defauts'))
        
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            print('==========>',file_path)
            file.save(file_path)

            fiichier = Fichier_charger(file.filename,current_user.id)
            db.session.add(fiichier)
            print('==========<',current_user.id)

            fichier = Fichier_charger.query.filter_by(nom=file.filename).first()
            user = session['_user_id']
            print('======GGGGGGG===================>',user)
            if fichier and user == current_user.id:
                if fichier:
                    flash(f'Le fichier {fichier.nom} est déja dans la base', 'success')
            else:

                if os.path.exists(file_path):
                    df = pd.read_excel(file_path)
                    print('==========>',df[['Libellé du Service (complet)', 'TYPE_ECHANT', 'Défaut (OUI/NON)']])

                    # Remplacer les valeurs 'nan' par une valeur par défaut ou une chaîne vide
                    df.replace({np.nan: None}, inplace=True)
                    # Itérer sur les lignes du dataframe et enregistrer dans la base de données
                    for _, row in df.iterrows():
                        file = Fichier(
                                numero_demande = row['N° Commande'],
                                enregistre_le  = row['Enregistré le'],         
                                date_resolution = row['Date de résolution'],            
                                libelle_service = row['Libellé du Service (complet)'],
                                saisi_par      = row['Saisi par'],
                                demandeur      = row['Demandeur'],
                                demandeur_entite = row['Demandeur : Entité (complète)'],
                                localisation    = row['Localisation (complète)'],
                                urgence_utilisateur = row['Urgence utilisateur'],
                                impact        = row['Impact'],
                                priorite    = row['Priorité'],
                                statut_demande = row['Statut de la demande'],
                                delai_resolution_hhmm = row ['Délai de résolution (hh:mm)'],
                                delai_resolution_min = row['Délai de résolution (min)'],
                                resolution_immediate = row['Resolution immédiate'],
                                resolu_par_groupe = row['Résolu par (groupe)'],
                                origine_demande = row['Date de résolution maximum'],
                                date_resolution_maximum = row['Description'],
                                description = row['Résolu par (intervenant)'],
                                resolu_par_intervenant = row['Résolu par (intervenant)'],
                                service_retard_hhmm = row['Service : Retard (hh:mm)'],
                                service_retard_min = row['Service : Retard (min)'],
                                group_fr = row['GROUP_FR'],
                                resolution = row['Résolution'],
                                sla = row['SLA'],
                                beneficiaire_courriel = row['Bénéficiaire : Courriel'],
                                xa_date_fin_de_mois = row['XA_DATE_FIN_DE_MOIS'],
                                xb_periode = row['XB_PERIODE'],
                                xc_statut_trait = row['XC_STATUT_TRAIT'],
                                xx_num_sequence = row['XX_NUM_SEQUENCE'],
                                xx_agent_transfert_dsi = row['XX_AGENT_TRNSFERT_DSI'],
                                xx_agent_responsable = row['XX_AGENT_RESPONSABLE'],
                                xx_service = row['XX_SERVICE'],
                                xx_intervalle_delai_res = row['XX_INTERVALLE_DELAI_RES'],
                                xx_delai30min =row['XX_DELAI30MIN'],
                                xx_delai1h = row['XX_DELAI1H'],
                                xx_delai2h = row['XX_DELAI2H'],
                                xx_delai1j = row['XX_DELAI1J'],
                                xx_delai2j = row['XX_DELAI2J'],
                                xx_respect_delais = row['XX_RESPECT_DELAIS'],
                                xx_retard_en_jours = row['XX_RETARD_EN_JOURS'],
                                xx_activite = row['XX_ACTYIVITE'],
                                xx_a_comptabiliser = row['XX_A_COMPTABILISER'],
                                xx_application = row['XX_APPLICATION'],
                                xx_dep_traitant = row['XX_DEP_TRAITANT'],
                                xx_direction = row['XX_DIRECTION'],
                                xx_agent_refus = row['XX_AGENT_REFUS'],
                                numero = row['N°'],
                                type_echant = row['TYPE_ECHANT'],
                                defaut = row['Défaut (OUI/NON)'],
                                type_description_defaut = row['Type'],
                                description_du_defaut = row['Description du Défaut'],
                                commentaires = row['Commentaires'],
                                note_defaut = row['NOTE_DEFAUT'],
                                agent_escalade = row['Agent ESCALADE'],
                                pertinence_escalade = row['Pertinence ESCALADE'],
                                type_erreur_escalade = row['TypeErreurEsacalade'],
                                actions_correctives_preventives = row['Actions Correctives/Préventives']
                            )

                        # db.session.add(file)
            db.session.commit()


            
            flash('Chargement des défauts effectué avec succès!', 'success')
            return redirect(url_for('chargement_defauts'))
        else:
            flash('Type de fichier non autorisé.', 'error')
            return redirect(url_for('chargement_defauts'))
    return render_template('chargement_defauts.html',fichiers=fichiers)


@app.route('/chargement-tickets', methods=['GET', 'POST'])
@login_required
def chargement_tickets():
    if request.method == 'POST':
        confirmation = request.form.get('confirmation')
        # noconf= request.form.get('noconf')

        if confirmation == 'Oui':
            file = request.files['file']
            liste_utile =[]
            if file.filename == '':
                flash('Aucun fichier sélectionné.', 'error')
                return redirect(url_for('chargement_tickets'))
            
            if file:
                try:
                    df = pd.read_excel(file)

                    # Nombre d'enregistrements rejetés
                    num_rejected = 0  
                    # Liste des enregistrements rejetés
                    rejected_records = []

                    df.fillna('', inplace=True)
                    for index, row in df.iterrows():
                        print('BBBBBBBBonjour')
                        # Récupérer les valeurs des colonnes
                        if pd.notnull(row['Date de résolution maximum']) or isinstance(row['Date de résolution maximum'], pd.Timestamp):
                            print('Date de résolution maximum')
                            utilisateur = User.query.filter(User.nom_abrege == row['XX_AGENT_RESPONSABLE'], User.state == 'Actif').first()
                            print('>>>>>>>>>>>>>>><<<<<<<<<<<<<<<<<<<<<<<<<', utilisateur)
                            if utilisateur:
                                print('Insertion successful')
                                # Charger l'enregistrement dans la base de données
                                ticket = Ticket(
                                    numero_demande=row['N° Commande'],
                                    enregistre_le=row['Enregistré le'],
                                    date_resolution=row['Date de résolution'],
                                    libelle_service=row['Libellé du Service (complet)'],
                                    demandeur=row['Demandeur'],
                                    statut_demande=row['Statut de la demande'],
                                    resolu_par=row['Résolu par (groupe)'],
                                    origine_demande=row['Origine de la demande'],
                                    date_resolution_max=row['Date de résolution maximum'],
                                    description=row['Description'],
                                    resolution=row['Résolution'],
                                    sla=row['SLA'],
                                    nom_abrege_agent=row['Bénéficiaire : Courriel'],
                                    type_echant=row['TYPE_ECHANT'],
                                    defaut=row['Défaut (OUI/NON)'],
                                    type_defaut=row['Type'],
                                    description_defaut=row['Description du Défaut'],
                                    commentaires_defaut=row['Commentaires'],
                                    periode=row['XB_PERIODE'],
                                    evaluateur=row['XX_AGENT_RESPONSABLE']
                                )
                                
                                db.session.add(ticket)
                            else:
                                # L'enregistrement est rejeté car l'utilisateur n'est pas déclaré ou actif
                                num_rejected += 1
                                rejected_records.append(row.to_dict())
                        else:
                            df.at[index, 'Date de résolution maximum'] = np.nan 

                    # Nombre d'enregistrements chargés avec succès
                    num_loaded = len(df) - num_rejected

                    flash(f"Nombre d'enregistrements à charger: {len(df)}", 'info')
                    flash(f"Nombre d'enregistrements chargés: {num_loaded}", 'success')
                    flash(f"Nombre d'enregistrements rejetés: {num_rejected}", 'warning')
                    # long = len(df)
                    # liste_utile.append(long,num_loaded,num_rejected)
                    
                    if num_rejected > 0:
                        folder_path = os.path.join(os.getcwd(), 'files_rejet')
                        os.makedirs(folder_path, exist_ok=True)

                        file_path = os.path.join(folder_path, f'rejected_records_{random.randint(12, 706)+random.randint(2, 6)}_{datetime.now()}.csv')
                        file_path = os.path.abspath(file_path)

                        field_names = ['N° de demande', 'Enregistré le', 'Date de résolution', 'Libellé du Service (complet)', 'Demandeur', 'Statut de la demande', 'Résolu par (groupe)', 'Origine de la demande', 'Date de résolution maximum', 'Description', 'Résolution', 'SLA', 'Bénéficiaire : Courriel', 'TYPE_ECHANT', 'Défaut (OUI/NON)', 'Type', 'Description du Défaut', 'Commentaires', 'XB_PERIODE', 'XX_AGENT_RESPONSABLE', 'reason', 'row_index']

                        with open(file_path, 'w', newline='') as csv_file:
                            fieldnames = rejected_records[0].keys() 
                            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                            writer.writeheader()
                            
                            writer.writerows(rejected_records)

                        # Chemin du répertoire d'upload

                    # db.session.commit()


                    flash("Le chargement des tickets a été validé avec succès.", 'success')
                except Exception as e:
                    flash(f"Une erreur s'est produite lors du chargement du fichier: {str(e)}", 'danger')
            else:
                flash("Aucun fichier sélectionné.", 'error')
        else:
            flash("Le chargement des tickets a été annulé.", 'warning')
        
        return redirect(url_for('chargement_tickets'))
    return render_template('chargement_tickets.html')

@app.route('/sonatel-gmec/details_tickets')
@login_required
def details_tickets():
    
    folder_path = os.path.join(os.getcwd(), 'files_rejet')

    # Obtenir la liste des fichiers dans le répertoire triés par date de modification
    file_list = glob.glob(os.path.join(folder_path, 'rejected_records_*.csv'))
    file_list.sort(key=os.path.getmtime)

    latest_file = file_list[-1]

    df = pd.read_csv(latest_file)
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10  # Nombre de lignes par page

    total = len(df)
    pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')

    # Obtention des données paginées
    start = (page - 1) * per_page
    end = start + per_page
    paginated_data = df.iloc[start:end]

    # print(df.head())
    return render_template('details_tickets.html',df=df,paginated_data=paginated_data, pagination=pagination)


@app.route('/traitement-ec', methods=['GET', 'POST'])
def traitement_ec():
    if request.method == 'POST':
        # Process defect treatment and validation
        flash('Traitement des EC effectué avec succès!', 'success')
        return redirect(url_for('traitement_ec'))
    return render_template('traitement_ec.html')




                                ########################################################
                                #                          API                         #
                                ########################################################


@app.route('/type_defaut')
@login_required
def type_defaut():
    defaut = TypeDefaut.query.all()

    return render_template('typedefaut.html',defaut=defaut)

@app.route('/api/data')
@login_required
def data():
    query = Fichier.query

    # search filter
    search = request.args.get('search')

    print('====>',request.args)
    if search:
        query = query.filter(db.or_(
            Fichier.demandeur.like(f'%{search}%'),
            Fichier.numero_demande.like(f'%{search}%')
        ))
    total = query.count()

    # sorting
    sort = request.args.get('sort')
    if sort:
        order = []
        for s in sort.split(','):
            direction = s[0]
            name = s[1:]
            if name not in ['numero_commande', 'enregistrer_le', 'date_resolution', 'libelle_service', 'statut_demande','defaut']:
                name = 'numero_commande'
            col = getattr(Fichier, name)
            if direction == '-':
                col = col.desc()
            order.append(col)
        if order:
            query = query.order_by(*order)

    # pagination
    start = request.args.get('start', type=int, default=-1)
    length = request.args.get('length', type=int, default=-1)
    if start != -1 and length != -1:
        query = query.offset(start).limit(length)

    # response
    return {
        'data': [user.to_dict() for user in query],
        'total': total,
    }


# @app.route('/type', methods=['GET', 'POST'])
# def type():
#     defauts = TypeDefaut.query.all()
#     return render_template('type.html',defauts=defauts)


# Main entry point
if __name__ == '__main__':
    app.run(debug=True)